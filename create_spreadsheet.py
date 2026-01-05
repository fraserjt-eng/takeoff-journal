#!/usr/bin/env python3
"""
High Performance Takeoff Tracker - Spreadsheet Generator
Creates a comprehensive Excel workbook for tracking daily performance
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from datetime import datetime, timedelta
import os

# Color definitions
COLORS = {
    'bg_primary': '0A1628',
    'bg_secondary': '1A2940',
    'accent_gold': 'D4A853',
    'accent_teal': '14B8A6',
    'accent_coral': 'F97316',
    'accent_purple': '8B5CF6',
    'text_primary': 'FFFFFF',
    'text_secondary': '94A3B8',
    'status_success': '22C55E',
    'status_warning': 'EAB308',
    'status_danger': 'EF4444',
    'domain_money': 'D4A853',
    'domain_health': '14B8A6',
    'domain_career': 'F97316',
    'domain_creative': '8B5CF6',
    'domain_love': 'EC4899',
    'domain_inner': '06B6D4',
}

def create_workbook():
    wb = Workbook()

    # Define styles
    header_font = Font(name='Arial', size=14, bold=True, color=COLORS['accent_gold'])
    title_font = Font(name='Arial', size=20, bold=True, color=COLORS['accent_gold'])
    subtitle_font = Font(name='Arial', size=12, color=COLORS['text_secondary'])
    normal_font = Font(name='Arial', size=11, color=COLORS['text_primary'])

    header_fill = PatternFill(start_color=COLORS['bg_secondary'], end_color=COLORS['bg_secondary'], fill_type='solid')
    bg_fill = PatternFill(start_color=COLORS['bg_primary'], end_color=COLORS['bg_primary'], fill_type='solid')
    gold_fill = PatternFill(start_color=COLORS['accent_gold'], end_color=COLORS['accent_gold'], fill_type='solid')

    thin_border = Border(
        left=Side(style='thin', color=COLORS['bg_secondary']),
        right=Side(style='thin', color=COLORS['bg_secondary']),
        top=Side(style='thin', color=COLORS['bg_secondary']),
        bottom=Side(style='thin', color=COLORS['bg_secondary'])
    )

    # Create tabs
    create_dashboard(wb, header_font, title_font, subtitle_font, normal_font, header_fill, bg_fill)
    create_daily_journal(wb, header_font, normal_font, header_fill, bg_fill)
    create_performance_analytics(wb, header_font, title_font, normal_font, header_fill, bg_fill)
    create_ai_insights(wb, header_font, title_font, normal_font, header_fill, bg_fill)
    create_vision_tracker(wb, header_font, title_font, normal_font, header_fill, bg_fill)
    create_weekly_review(wb, header_font, title_font, normal_font, header_fill, bg_fill)
    create_monthly_review(wb, header_font, title_font, normal_font, header_fill, bg_fill)
    create_settings(wb, header_font, title_font, normal_font, header_fill, bg_fill)

    return wb

def create_dashboard(wb, header_font, title_font, subtitle_font, normal_font, header_fill, bg_fill):
    ws = wb.active
    ws.title = "Executive Dashboard"

    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 5
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 45

    # Title
    ws['B2'] = "HIGH PERFORMANCE TAKEOFF TRACKER"
    ws['B2'].font = title_font
    ws.merge_cells('B2:D2')

    ws['B3'] = '"The becoming is the point." - J. Fraser'
    ws['B3'].font = subtitle_font
    ws.merge_cells('B3:D3')

    # Date info
    ws['B4'] = f"Date: {datetime.now().strftime('%B %d, %Y')}"
    ws['B4'].font = normal_font
    ws['C4'] = f"Week: {datetime.now().isocalendar()[1]}"
    ws['C4'].font = normal_font
    ws['D4'] = f"Q{(datetime.now().month-1)//3 + 1}"
    ws['D4'].font = normal_font

    # Performance Snapshot Section
    ws['B6'] = "PERFORMANCE SNAPSHOT"
    ws['B6'].font = header_font
    ws.merge_cells('B6:D6')

    metrics = [
        ("Daily Performance Score", "='Performance Analytics'!C5", "%"),
        ("Energy Level", "='Daily Journal'!AA2", "/10"),
        ("Weekly Goals Progress", "=COUNTIF('Daily Journal'!AF2:AF8,\"Y\")/7*100", "%"),
        ("Leadership Impact", "='Daily Journal'!AD2", "/10"),
        ("Work-Life Balance", "=(100-ABS('Daily Journal'!AS2-70))", "%"),
        ("Trend", "=IF('Performance Analytics'!C5>'Performance Analytics'!C6,\"Rising\",IF('Performance Analytics'!C5<'Performance Analytics'!C6,\"Falling\",\"Stable\"))", ""),
    ]

    for i, (label, formula, suffix) in enumerate(metrics, start=7):
        ws[f'B{i}'] = label
        ws[f'B{i}'].font = normal_font
        ws[f'C{i}'] = formula
        ws[f'C{i}'].font = Font(name='Arial', size=11, bold=True, color=COLORS['accent_gold'])
        ws[f'D{i}'] = suffix
        ws[f'D{i}'].font = normal_font

    # AI Insights Section
    ws['F6'] = "AI INSIGHTS & RECOMMENDATIONS"
    ws['F6'].font = header_font
    ws.merge_cells('F6:G6')

    insights = [
        "Performance Optimization: Review and apply insights from 'AI Insights' tab",
        "Energy Management: Track patterns in your energy scores",
        "Focus Areas: Identify lowest-scoring habits for improvement",
        "Consistency Check: Maintain daily journal entries for best insights",
        "Weekly Review: Complete Sunday reviews for pattern recognition",
        "Path Balance: Monitor 70/30 split between Path A and Path B",
    ]

    for i, insight in enumerate(insights, start=7):
        ws[f'F{i}'] = f"{i-6}."
        ws[f'F{i}'].font = normal_font
        ws[f'G{i}'] = insight
        ws[f'G{i}'].font = normal_font

    # Weekly Summary Section
    ws['B15'] = "WEEKLY HABIT SCORES"
    ws['B15'].font = header_font

    days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    habits = ['Clarity', 'Energy', 'Necessity', 'Productivity', 'Influence', 'Courage']

    # Headers
    for j, day in enumerate(days, start=3):
        ws.cell(row=16, column=j, value=day)
        ws.cell(row=16, column=j).font = Font(name='Arial', size=10, bold=True, color=COLORS['accent_teal'])

    # Habit rows
    for i, habit in enumerate(habits, start=17):
        ws.cell(row=i, column=2, value=habit)
        ws.cell(row=i, column=2).font = normal_font
        for j in range(3, 10):
            # Reference to Daily Journal data (would need actual row references)
            ws.cell(row=i, column=j, value=5)  # Placeholder
            ws.cell(row=i, column=j).font = normal_font

    # Apply background color to used range
    for row in ws.iter_rows(min_row=1, max_row=25, min_col=1, max_col=10):
        for cell in row:
            cell.fill = bg_fill

def create_daily_journal(wb, header_font, normal_font, header_fill, bg_fill):
    ws = wb.create_sheet("Daily Journal")

    # Define all columns
    columns = [
        ('A', 'Date', 12),
        ('B', 'Day', 10),
        ('C', 'Week', 8),
        ('D', 'Quarter', 8),
        ('E', 'Today\'s Message', 40),
        ('F', 'Goal 1', 25),
        ('G', 'Goal 2', 25),
        ('H', 'Goal 3', 25),
        ('I', 'Task 1', 20),
        ('J', 'Task 1 Done', 10),
        ('K', 'Task 2', 20),
        ('L', 'Task 2 Done', 10),
        ('M', 'Task 3', 20),
        ('N', 'Task 3 Done', 10),
        ('O', 'Reach Out 1', 15),
        ('P', 'Reach Out 2', 15),
        ('Q', 'Reach Out 3', 15),
        ('R', 'Prompt 1', 30),
        ('S', 'Prompt 2', 30),
        ('T', 'Prompt 3', 30),
        ('U', 'Prompt 4', 30),
        ('V', 'Prompt 5', 30),
        ('W', 'Prompt 6', 30),
        ('X', 'Prompt 7', 30),
        ('Y', 'Prompt 8', 30),
        ('Z', 'Prompt 9', 30),
        ('AA', 'Prompt 10', 30),
        ('AB', 'Prompt 11', 30),
        ('AC', 'Clarity', 10),
        ('AD', 'Energy', 10),
        ('AE', 'Necessity', 10),
        ('AF', 'Productivity', 10),
        ('AG', 'Influence', 10),
        ('AH', 'Courage', 10),
        ('AI', 'Overall Score', 12),
        ('AJ', 'Money Progress', 12),
        ('AK', 'Health Progress', 12),
        ('AL', 'Career Progress', 12),
        ('AM', 'Creative Progress', 12),
        ('AN', 'Love Progress', 12),
        ('AO', 'Inner Progress', 12),
        ('AP', 'Money Notes', 25),
        ('AQ', 'Health Notes', 25),
        ('AR', 'Career Notes', 25),
        ('AS', 'Creative Notes', 25),
        ('AT', 'Love Notes', 25),
        ('AU', 'Inner Notes', 25),
        ('AV', 'Path A %', 10),
        ('AW', 'Win 1', 30),
        ('AX', 'Win 2', 30),
        ('AY', 'Win 3', 30),
        ('AZ', 'Improvement', 30),
        ('BA', 'Gratitude', 30),
        ('BB', 'Tomorrow Priority', 30),
        ('BC', 'Timestamp', 20),
    ]

    # Set headers and column widths
    for col_letter, header, width in columns:
        col_idx = ord(col_letter[0]) - ord('A') + 1
        if len(col_letter) > 1:
            col_idx = (ord(col_letter[0]) - ord('A') + 1) * 26 + (ord(col_letter[1]) - ord('A') + 1)

        ws[f'{col_letter}1'] = header
        ws[f'{col_letter}1'].font = header_font
        ws[f'{col_letter}1'].fill = header_fill
        ws.column_dimensions[col_letter].width = width

    # Add formula for Overall Score column (AI)
    ws['AI2'] = '=AVERAGE(AC2:AH2)'

    # Add sample row with today's date
    ws['A2'] = datetime.now().strftime('%Y-%m-%d')
    ws['B2'] = datetime.now().strftime('%A')
    ws['C2'] = datetime.now().isocalendar()[1]
    ws['D2'] = f"Q{(datetime.now().month-1)//3 + 1}"

    # Freeze header row
    ws.freeze_panes = 'A2'

def create_performance_analytics(wb, header_font, title_font, normal_font, header_fill, bg_fill):
    ws = wb.create_sheet("Performance Analytics")

    # Title
    ws['B2'] = "PERFORMANCE ANALYTICS"
    ws['B2'].font = title_font
    ws.merge_cells('B2:H2')

    # Set column widths
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 15
    ws.column_dimensions['B'].width = 20

    # Performance Trends Header
    ws['B4'] = "PERFORMANCE TRENDS (Last 30 Days)"
    ws['B4'].font = header_font

    # Column headers
    headers = ['Date', 'Clarity', 'Energy', 'Necessity', 'Productivity', 'Influence', 'Courage', 'Overall']
    for i, header in enumerate(headers, start=2):
        ws.cell(row=5, column=i, value=header)
        ws.cell(row=5, column=i).font = Font(name='Arial', size=11, bold=True, color=COLORS['accent_teal'])
        ws.cell(row=5, column=i).fill = header_fill

    # Sample data rows (referencing Daily Journal)
    for row in range(6, 36):
        ws.cell(row=row, column=2, value=f"='Daily Journal'!A{row-4}")
        for col in range(3, 10):
            # These would reference actual Daily Journal columns
            ws.cell(row=row, column=col, value="")

    # Key Metrics Section
    ws['B38'] = "KEY METRICS"
    ws['B38'].font = header_font

    metrics = [
        ('Average Daily Score:', '=AVERAGE(I6:I35)'),
        ('Best Performing Habit:', '=INDEX(C5:H5,MATCH(MAX(AVERAGE(C6:C35),AVERAGE(D6:D35),AVERAGE(E6:E35),AVERAGE(F6:F35),AVERAGE(G6:G35),AVERAGE(H6:H35)),{AVERAGE(C6:C35),AVERAGE(D6:D35),AVERAGE(E6:E35),AVERAGE(F6:F35),AVERAGE(G6:G35),AVERAGE(H6:H35)},0))'),
        ('Improvement Opportunity:', '=INDEX(C5:H5,MATCH(MIN(AVERAGE(C6:C35),AVERAGE(D6:D35),AVERAGE(E6:E35),AVERAGE(F6:F35),AVERAGE(G6:G35),AVERAGE(H6:H35)),{AVERAGE(C6:C35),AVERAGE(D6:D35),AVERAGE(E6:E35),AVERAGE(F6:F35),AVERAGE(G6:G35),AVERAGE(H6:H35)},0))'),
        ('Consistency Score:', '=COUNTA(B6:B35)/30*100'),
        ('Weekly Trend:', '=IF(AVERAGE(I6:I12)>AVERAGE(I13:I19),"Improving","Needs Focus")'),
    ]

    for i, (label, formula) in enumerate(metrics, start=39):
        ws.cell(row=i, column=2, value=label)
        ws.cell(row=i, column=2).font = normal_font
        ws.cell(row=i, column=3, value=formula)
        ws.cell(row=i, column=3).font = Font(name='Arial', size=11, bold=True, color=COLORS['accent_gold'])

    # Trend Analysis Section
    ws['B46'] = "TREND ANALYSIS"
    ws['B46'].font = header_font

    ws['B47'] = "Energy vs Productivity Correlation:"
    ws['C47'] = '=CORREL(D6:D35,F6:F35)'
    ws['B48'] = "Clarity vs Overall Correlation:"
    ws['C48'] = '=CORREL(C6:C35,I6:I35)'

def create_ai_insights(wb, header_font, title_font, normal_font, header_fill, bg_fill):
    ws = wb.create_sheet("AI Insights")

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 60

    # Title
    ws['B2'] = "AI-POWERED INSIGHTS"
    ws['B2'].font = title_font
    ws.merge_cells('B2:C2')

    # Personalized Recommendations
    ws['B4'] = "PERSONALIZED RECOMMENDATIONS"
    ws['B4'].font = header_font

    recommendations = [
        ("Performance Optimization", '=IF(\'Performance Analytics\'!C39<7,"Focus on improving your weakest habit area","Maintain your strong performance patterns")'),
        ("Stress Management", '=IF(\'Daily Journal\'!AD2<5,"Consider energy management techniques - your energy score is low","Your energy levels are healthy - maintain current practices")'),
        ("Leadership Effectiveness", '=IF(\'Daily Journal\'!AG2<6,"Look for opportunities to influence and lead others","Your leadership impact is strong - continue mentoring")'),
        ("Work-Life Balance", '=IF(ABS(\'Daily Journal\'!AV2-70)>20,"Adjust your Path A/B balance toward the 70/30 target","Good balance between creative work and stability")'),
        ("Energy Management", '=IF(\'Daily Journal\'!AD2<\'Daily Journal\'!AC2,"Energy is lagging behind clarity - focus on physical wellness","Energy and clarity are well-aligned")'),
        ("Goal Achievement", '=IF(COUNTBLANK(\'Daily Journal\'!F2:H2)>0,"Set clear daily goals to improve focus","Goals are set - focus on execution")'),
    ]

    for i, (category, formula) in enumerate(recommendations, start=5):
        ws.cell(row=i, column=2, value=category)
        ws.cell(row=i, column=2).font = Font(name='Arial', size=11, bold=True, color=COLORS['accent_teal'])
        ws.cell(row=i, column=3, value=formula)
        ws.cell(row=i, column=3).font = normal_font

    # Pattern Recognition
    ws['B13'] = "PATTERN RECOGNITION"
    ws['B13'].font = header_font

    patterns = [
        ("Best Performance Days", "Track your highest-scoring days to identify optimal conditions"),
        ("Energy Peak Times", "Note when your energy scores are highest"),
        ("Stress Triggers", "Identify situations that correlate with low scores"),
        ("Productivity Patterns", "Find which activities lead to highest productivity"),
        ("Leadership Moments", "Document when you have the most influence"),
    ]

    for i, (pattern, description) in enumerate(patterns, start=14):
        ws.cell(row=i, column=2, value=pattern)
        ws.cell(row=i, column=2).font = Font(name='Arial', size=11, bold=True, color=COLORS['accent_coral'])
        ws.cell(row=i, column=3, value=description)
        ws.cell(row=i, column=3).font = normal_font

    # Predictive Analytics
    ws['B21'] = "PREDICTIVE ANALYTICS"
    ws['B21'].font = header_font

    ws['B22'] = "Based on your patterns, tomorrow you should:"
    ws['B22'].font = normal_font

    predictions = [
        '=IF(\'Daily Journal\'!AD2<6,"1. Prioritize rest and recovery activities","1. Channel your high energy into bold action"))',
        '=IF(\'Daily Journal\'!AC2<6,"2. Start with a clarity-building morning routine","2. Dive into your most important creative work"))',
        '=IF(\'Daily Journal\'!AG2<6,"3. Schedule time for meaningful connections","3. Lead a meeting or mentor someone"))',
    ]

    for i, formula in enumerate(predictions, start=23):
        ws.cell(row=i, column=2, value=formula)
        ws.cell(row=i, column=2).font = normal_font

def create_vision_tracker(wb, header_font, title_font, normal_font, header_fill, bg_fill):
    ws = wb.create_sheet("Vision Board Tracker")

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 35

    # Title
    ws['B2'] = "2026 VISION BOARD TRACKER"
    ws['B2'].font = title_font
    ws.merge_cells('B2:G2')

    # Headers
    headers = ['Domain', 'Target', 'Current', 'Progress %', 'Last Updated', 'Notes']
    for i, header in enumerate(headers, start=2):
        ws.cell(row=4, column=i, value=header)
        ws.cell(row=4, column=i).font = Font(name='Arial', size=11, bold=True, color=COLORS['accent_gold'])
        ws.cell(row=4, column=i).fill = header_fill

    # Domain data
    domains = [
        ('Money & Finance', '$15K/month revenue', '$0', '0%', '', '5 -> 10 -> 20 clients'),
        ('Health & Fitness', '3x/week, 185-197 lbs', '0x/week', '0%', '', 'Strength training focus'),
        ('Career/BCCS', 'Excellence + Legacy', 'In Progress', '0%', '', 'Build automation systems'),
        ('Creative Ventures', '4 books, 5K followers', '0 books', '0%', '', 'SHIP IT! Path A focus'),
        ('Love & Relationship', '12 date nights', '0 dates', '0%', '', 'Presence over performance'),
        ('Inner Peace', 'Daily practice', '0%', '0%', '', 'The spiral continues'),
    ]

    domain_colors = [
        COLORS['domain_money'],
        COLORS['domain_health'],
        COLORS['domain_career'],
        COLORS['domain_creative'],
        COLORS['domain_love'],
        COLORS['domain_inner'],
    ]

    for i, (domain, target, current, progress, updated, notes) in enumerate(domains, start=5):
        ws.cell(row=i, column=2, value=domain)
        ws.cell(row=i, column=2).font = Font(name='Arial', size=11, bold=True, color=domain_colors[i-5])
        ws.cell(row=i, column=3, value=target)
        ws.cell(row=i, column=4, value=current)
        ws.cell(row=i, column=5, value=progress)
        ws.cell(row=i, column=6, value=updated)
        ws.cell(row=i, column=7, value=notes)

    # Path Allocation Section
    ws['B13'] = "PATH ALLOCATION TRACKING"
    ws['B13'].font = header_font

    ws['B14'] = "Weekly Average - Path A (SHIP IT!):"
    ws['C14'] = '=AVERAGE(\'Daily Journal\'!AV2:AV8)'
    ws['D14'] = "%"

    ws['B15'] = "Weekly Average - Path B (Stability):"
    ws['C15'] = '=100-C14'
    ws['D15'] = "%"

    ws['B16'] = "Target Split:"
    ws['C16'] = "70% / 30%"

    ws['B17'] = "Variance from Target:"
    ws['C17'] = '=ABS(C14-70)'
    ws['D17'] = "% off"

def create_weekly_review(wb, header_font, title_font, normal_font, header_fill, bg_fill):
    ws = wb.create_sheet("Weekly Review")

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15

    # Title
    ws['B2'] = "WEEKLY REVIEW"
    ws['B2'].font = title_font

    # Week selector
    ws['B4'] = "Week Number:"
    ws['C4'] = datetime.now().isocalendar()[1]
    ws['C4'].font = Font(name='Arial', size=14, bold=True, color=COLORS['accent_gold'])

    # Weekly Wins
    ws['B6'] = "WEEKLY WINS (Top 3)"
    ws['B6'].font = header_font

    for i in range(7, 10):
        ws.cell(row=i, column=2, value=f"{i-6}.")
        ws.cell(row=i, column=3, value="")
        ws.merge_cells(f'C{i}:E{i}')

    # Weekly Challenges
    ws['B11'] = "WEEKLY CHALLENGES"
    ws['B11'].font = header_font
    ws['C12'] = ""
    ws.merge_cells('C12:E12')

    # Habit Score Averages
    ws['B14'] = "HABIT SCORE AVERAGES"
    ws['B14'].font = header_font

    # Day headers
    days = ['', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun', 'Avg']
    for i, day in enumerate(days, start=2):
        ws.cell(row=15, column=i, value=day)
        ws.cell(row=15, column=i).font = Font(name='Arial', size=10, bold=True, color=COLORS['accent_teal'])

    # Habit rows
    habits = ['Clarity', 'Energy', 'Necessity', 'Productivity', 'Influence', 'Courage', 'Overall']
    for i, habit in enumerate(habits, start=16):
        ws.cell(row=i, column=2, value=habit)
        ws.cell(row=i, column=2).font = normal_font
        # Average formula in last column
        ws.cell(row=i, column=10, value=f'=AVERAGE(C{i}:I{i})')

    # Insights Section
    ws['B25'] = "INSIGHTS FOR NEXT WEEK"
    ws['B25'].font = header_font
    ws['C26'] = '=IF(J22<7,"Focus on improving overall performance","Maintain momentum and push for excellence")'

    # Commitment
    ws['B28'] = "COMMITMENT FOR NEXT WEEK"
    ws['B28'].font = header_font
    ws['C29'] = ""
    ws.merge_cells('C29:F29')

def create_monthly_review(wb, header_font, title_font, normal_font, header_fill, bg_fill):
    ws = wb.create_sheet("Monthly Review")

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20

    # Title
    ws['B2'] = "MONTHLY REVIEW"
    ws['B2'].font = title_font

    # Month selector
    ws['B4'] = "Month:"
    ws['C4'] = datetime.now().strftime('%B %Y')
    ws['C4'].font = Font(name='Arial', size=14, bold=True, color=COLORS['accent_gold'])

    # Monthly Scorecard
    ws['B6'] = "MONTHLY SCORECARD"
    ws['B6'].font = header_font

    scorecard_items = [
        ('Average Daily Performance:', '=AVERAGE(\'Daily Journal\'!AI2:AI32)', '%'),
        ('Consistency (days completed):', '=COUNTA(\'Daily Journal\'!A2:A32)', '/31'),
        ('Best Week:', 'Week 1', ''),
        ('Improvement from Last Month:', '+0', '%'),
    ]

    for i, (label, value, suffix) in enumerate(scorecard_items, start=7):
        ws.cell(row=i, column=2, value=label)
        ws.cell(row=i, column=2).font = normal_font
        ws.cell(row=i, column=3, value=value)
        ws.cell(row=i, column=3).font = Font(name='Arial', size=11, bold=True, color=COLORS['accent_gold'])
        ws.cell(row=i, column=4, value=suffix)

    # Domain Progress
    ws['B13'] = "DOMAIN PROGRESS"
    ws['B13'].font = header_font

    domains = ['Money & Finance', 'Health & Fitness', 'Career/BCCS', 'Creative Ventures', 'Love & Relationship', 'Inner Peace']
    for i, domain in enumerate(domains, start=14):
        ws.cell(row=i, column=2, value=domain)
        ws.cell(row=i, column=2).font = normal_font
        ws.cell(row=i, column=3, value="0%")
        ws.cell(row=i, column=4, value="of monthly target")

    # Wins of the Month
    ws['B22'] = "WINS OF THE MONTH"
    ws['B22'].font = header_font
    for i in range(23, 26):
        ws.cell(row=i, column=2, value=f"{i-22}.")
        ws.cell(row=i, column=3, value="")

    # Focus Areas
    ws['B28'] = "FOCUS AREAS FOR NEXT MONTH"
    ws['B28'].font = header_font
    ws['C29'] = '=\'AI Insights\'!C6'  # Reference to lowest scoring habit

def create_settings(wb, header_font, title_font, normal_font, header_fill, bg_fill):
    ws = wb.create_sheet("Settings & Reference")

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40

    # Title
    ws['B2'] = "SETTINGS & REFERENCE"
    ws['B2'].font = title_font

    # User Settings
    ws['B4'] = "USER SETTINGS"
    ws['B4'].font = header_font

    settings = [
        ('Name:', 'Joshua Fraser, Ed.D.'),
        ('Start Date:', 'January 1, 2026'),
        ('Time Zone:', 'America/Chicago'),
    ]

    for i, (label, value) in enumerate(settings, start=5):
        ws.cell(row=i, column=2, value=label)
        ws.cell(row=i, column=2).font = normal_font
        ws.cell(row=i, column=3, value=value)
        ws.cell(row=i, column=3).font = Font(name='Arial', size=11, color=COLORS['accent_gold'])

    # Scoring Reference
    ws['B10'] = "SCORING REFERENCE"
    ws['B10'].font = header_font

    scores = [
        ('1-4:', 'Needs Work (Red)'),
        ('5-6:', 'Making Progress (Yellow)'),
        ('7-8:', 'On Track (Green)'),
        ('9-10:', 'Exceptional (Gold)'),
    ]

    score_colors = [COLORS['status_danger'], COLORS['status_warning'], COLORS['status_success'], COLORS['accent_gold']]

    for i, ((range_label, description), color) in enumerate(zip(scores, score_colors), start=11):
        ws.cell(row=i, column=2, value=range_label)
        ws.cell(row=i, column=2).font = Font(name='Arial', size=11, bold=True, color=color)
        ws.cell(row=i, column=3, value=description)
        ws.cell(row=i, column=3).font = normal_font

    # Domain Definitions
    ws['B17'] = "DOMAIN DEFINITIONS"
    ws['B17'].font = header_font

    domains = [
        ('Money & Finance:', '$15K/mo revenue, 20 clients, debt freedom'),
        ('Health & Fitness:', '3x/week strength training, 185-197 lbs'),
        ('Career/BCCS:', 'Excellence while transitioning, legacy systems'),
        ('Creative Ventures:', '4 books shipped, 5K followers, 2K subscribers'),
        ('Love & Relationship:', '12 date nights, presence over performance'),
        ('Inner Peace:', 'Daily practice, the spiral continues'),
    ]

    for i, (domain, definition) in enumerate(domains, start=18):
        ws.cell(row=i, column=2, value=domain)
        ws.cell(row=i, column=2).font = Font(name='Arial', size=11, bold=True, color=COLORS['accent_teal'])
        ws.cell(row=i, column=3, value=definition)
        ws.cell(row=i, column=3).font = normal_font

    # Path Allocation
    ws['B26'] = "PATH ALLOCATION"
    ws['B26'].font = header_font

    ws['B27'] = "Path A (SHIP IT!):"
    ws['C27'] = "70% - Creative ventures, platform, books"
    ws['B28'] = "Path B (Stability):"
    ws['C28'] = "30% - BCCS excellence, family, health"

# Main execution
if __name__ == "__main__":
    print("Creating High Performance Takeoff Tracker...")
    wb = create_workbook()

    output_path = "/Users/jfraser/Desktop/TakeoffSystem/High_Performance_Takeoff_Tracker.xlsx"
    wb.save(output_path)
    print(f"Spreadsheet saved to: {output_path}")
    print("Done!")
